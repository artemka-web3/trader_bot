import aiohttp
import pandas as pd
import asyncio
import requests as re
from datetime import datetime, timedelta
import datetime as dt
import time
import logging
from openpyxl import load_workbook

offset = dt.timezone(timedelta(hours=3))

# Create an instance of asyncio event loop

#login – email, указанный при регистрации на сайте moex.com
login = 'kazakovoleg797@gmail.com'
#password – пароль от учетной записи на сайте moex.com
password = "Inkgroup12!"
s = re.Session()
s.get('https://passport.moex.com/authenticate', auth=(login, password))
headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:66.0) Gecko/20100101 Firefox/66.0"}
cookies = {'MicexPassportCert': s.cookies['MicexPassportCert']}
s.close()
workbook = load_workbook(filename='shares.xlsx')

def get_value_by_ticker(ticker):
    sheet = workbook.active
    row_counter = 2
    while row_counter < 249:
        cell = sheet.cell(row = row_counter, column = 2)
        if cell.value == ticker:
            name_cell = sheet.cell(row = row_counter, column = 3)
            return name_cell.value

# GET ONE STOCK DATA
async def fetch_stock(session, url, headers, cookies):
    async with session.get(url, headers=headers, cookies=cookies) as response:
        return await response.json()

async def one_stock(url, headers, cookies):
    async with aiohttp.ClientSession() as session:
        data = await fetch_stock(session, url, headers, cookies)
        name = data['securities']['data'][0][9]
        return (
            data['securities']['data'][0][0], # SECID, 
            name, # SECNAME
            data['securities']['data'][0][4], # LOTSIZE
            data['marketdata']['data'][0][20], # DAY CHANGE %
        )

    
async def get_stock_data(security):
    url_get_sec = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{security}.json"
    return await one_stock(url_get_sec, headers, cookies)

# GET ALL SECURITIES
url_get_secs = 'https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities.json'
async def fetch_all_securities(session, url, headers, cookies):
    async with session.get(url, headers=headers, cookies=cookies) as response:
        return await response.json()

async def all_securities(url, headers, cookies):
    async with aiohttp.ClientSession() as session:
        data = await fetch_all_securities(session, url, headers, cookies)
        return data['securities']['data']
    
async def all_marketdata(url, headers, cookies):
    async with aiohttp.ClientSession() as session:
        data = await fetch_all_securities(session, url, headers, cookies)
        return data['marketdata']['data']
    
async def get_securities():
    return await all_securities(url_get_secs, headers, cookies)

async def get_marketdata():
    return await all_marketdata(url_get_secs, headers, cookies)

# TRACK CURRENT VOLUME
async def fetch_current_volume(session, url, headers, cookies):
    async with session.get(url, headers=headers, cookies=cookies) as response:
        return await response.json()

async def get_current_volume(url, headers, cookies):
    async with aiohttp.ClientSession() as session:
        data = await fetch_current_volume(session, url, headers, cookies)
        return data
    
async def get_current_stock_volume(security):
    login = 'kazakovoleg797@gmail.com'
    #password – пароль от учетной записи на сайте moex.com
    password = "Inkgroup12!"
    s = re.Session()
    s.get('https://passport.moex.com/authenticate', auth=(login, password))
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:66.0) Gecko/20100101 Firefox/66.0"}
    cookies = {'MicexPassportCert': s.cookies['MicexPassportCert']}
    s.close()
    current_date = datetime.now(offset)  - timedelta(seconds=60) # - timedelta(hours=10)
    cur_time = ("0" +str(current_date.hour) if len(str(current_date.hour)) < 2 else str(current_date.hour)) + ":" + ("0" +str(current_date.minute) if len(str(current_date.minute)) < 2 else str(current_date.minute))
    today = current_date.strftime('%Y-%m-%d')
    start_from_for_today = 0
    while True:
        url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{security}/candles.json?from={today}&till={today}&interval=1&start={start_from_for_today}"
        cur_data= await get_current_volume(url, headers, cookies)
        url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{security}/candles.json?from={today}&till={today}&interval=1&start={start_from_for_today}"
        cur_data = await get_current_volume(url, headers, cookies)
        if len(cur_data['candles']['data']) == 0:
            return -200
        else: 
            for candle_data in cur_data['candles']['data']:
                if cur_time in candle_data[6]:
                    print("CUR: ", candle_data)
                    return candle_data
                else:
                    start_from_for_today += 1
# TRACK PREV MINUTE TIME
async def fetch_prevmin_price(session, url, headers, cookies):
    async with session.get(url, headers=headers, cookies=cookies) as response:
        return await response.json()

async def get_prevmin_price(url, headers, cookies):
    async with aiohttp.ClientSession() as session:
        data = await fetch_prevmin_price(session, url, headers, cookies)
        return data
    
async def get_prevmin_stock_price(security):
    login = 'kazakovoleg797@gmail.com'
    #password – пароль от учетной записи на сайте moex.com
    password = "Inkgroup12!"
    s = re.Session()
    s.get('https://passport.moex.com/authenticate', auth=(login, password))
    headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.14; rv:66.0) Gecko/20100101 Firefox/66.0"}
    cookies = {'MicexPassportCert': s.cookies['MicexPassportCert']}
    s.close()
    prevmin_date = datetime.now(offset) - timedelta(seconds=120) # - timedelta(hours=10)
    prevmin_time = ("0" +str(prevmin_date.hour) if len(str(prevmin_date.hour)) < 2 else str(prevmin_date.hour)) + ":" + ("0" +str(prevmin_date.minute) if len(str(prevmin_date.minute)) < 2 else str(prevmin_date.minute))
    today = prevmin_date.strftime('%Y-%m-%d')
    start_from_for_today = 0
    while True:
        url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{security}/candles.json?from={today}&till={today}&interval=1&start={start_from_for_today}"
        prev_minute_data= await get_prevmin_price(url, headers, cookies)
        if len(prev_minute_data['candles']['data']) == 0:
            return -200
        else: 
            for candle_data in prev_minute_data['candles']['data']:
                if prevmin_time in candle_data[6]:
                    print("PREV: ", candle_data)
                    return candle_data
                else:
                    start_from_for_today += 1


# PRICE CHANGE
async def get_price_change(security):
    current_candle = await get_current_stock_volume(security)
    prev_candle = await get_prevmin_stock_price(security)
    if prev_candle == -200 or current_candle == -200:
        raise Exception(f"\n{current_candle}\n{prev_candle}\nPRICE CHANGE COUNTING ERROR")
    current_close = current_candle[1]
    prev_close = prev_candle[1]
    #return current_candle, current_close, current_close, prev_close
    return round((float(current_close) * 100 / float(prev_close)) - 100, 3)


# GET ALL MINUTE VOLUMES WITHIN PAST 7 DAYS
async def fetch_prev(session, url, headers, cookies):
    async with session.get(url, headers=headers, cookies=cookies) as response:
        return await response.json()

async def get_prev(url, headers, cookies):
    async with aiohttp.ClientSession() as session:
        data = await fetch_prev(session, url, headers, cookies)
        return data
    
async def get_prev_avg_volume(volumes_dict):
    secs = await get_securities()
    for sec in secs:
        counter = 1
        empty_days = 0
        minutes = 0
        value = 0
        print(sec[0])
        volumes_dict[sec[0]] = 0
        while counter < 2:
            prev_date = (datetime.now(offset)- timedelta(days=counter)).strftime('%Y-%m-%d') # - timedelta(hours=10)
            url_hour = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{sec[0]}/candles.json?from={prev_date}&till={prev_date}&interval=60&start=0"
            prev_data_hour = await get_prev(url_hour, headers, cookies)
            if len(prev_data_hour['candles']['data']) != 0:  
                url_day = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{sec[0]}/candles.json?from={prev_date}&till={prev_date}&interval=24&start=0"
                prev_data_day = await get_prev(url_day, headers, cookies)
                for i in prev_data_day['candles']['data']:
                    if '23:' in i[7]:
                        minutes = 840
                    elif '18:' in i[7]:
                        minutes = 540
                value = prev_data_day['candles']['data'][0][4]
            counter += 1
        if value != 0:
            volumes_dict[sec[0]] += round(value / minutes, 6)
            volumes_dict[sec[0]] = volumes_dict[sec[0]] / (counter - 1)
        else:
            volumes_dict[sec[0]] = f'Ошибка получения данных об акции {sec[0]}'

        print(volumes_dict[sec[0]])
    return volumes_dict

# loop = asyncio.get_event_loop()
# print(loop.run_until_complete(get_stock_data('ABRD')))

# async def get_prev_avg_volume():
#     global volumes_dict
#     secs = await get_securities()
#     for sec in secs:
#         current_date = datetime.now(offset) # - timedelta(hours=10)
#         volumes = []
#         empty_days_counter = 0
#         counter = 1
#         while counter < 2:
#             start = 1
#             prev_date = current_date - timedelta(days=counter+empty_days_counter)
#             prev_date_str = str(prev_date.strftime('%Y-%m-%d'))
#             while True:
#                 try:
#                     url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{sec[0]}/candles.json?from={prev_date_str}&till={prev_date_str}&interval=1&start={start-1}"
#                     prev_data = await get_prev(url, headers=headers, cookies=cookies)
#                     if len(prev_data['candles']['data']) == 0:
#                         empty_days_counter += 1
#                         break
#                     else:
#                         if len(prev_data['candles']['data']) == 0 or counter == 2 or not prev_data['candles']['data'][0][4]:
#                             break
#                         url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{sec[0]}/candles.json?from={prev_date_str}&till={prev_date_str}&interval=1&start={start-1}"
#                         prev_data = await get_prev(url, headers=headers, cookies=cookies)
#                         print(prev_data['candles']['data'][0])
#                         volumes.append(float(prev_data['candles']['data'][0][4]))
#                         print(f"{prev_data['candles']['data'][0][4]}")
#                         print(f"START: {start-1}")
#                         print(prev_date_str)
#                         await asyncio.sleep(0.2)
#                     start += 1
#                 except Exception as e:
#                     print(e)
#                     await asyncio.sleep(5)
#                     continue
#             counter += 1
#         volumes_dict[f'{sec[0]}'] = sum(volumes) / start
#     return volumes_dict
# async def get_prev_avg_volume():
#     secs = await get_securities()
#     for sec in secs:
#         current_date = datetime.now(offset) # - timedelta(hours=10)
#         volumes = []
#         empty_days_counter = 0
#         counter = 1
#         while counter < 2:
#             start = 1
#             prev_date = current_date - timedelta(days=counter+empty_days_counter)
#             prev_date_str = str(prev_date.strftime('%Y-%m-%d'))
#             url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{sec[0]}/candles.json?from={prev_date_str}&till={prev_date_str}&interval=1&start={start-1}"
#             prev_data = await get_prev(url, headers=headers, cookies=cookies)
#             if len(prev_data['candles']['data']) == 0:
#                 empty_days_counter += 1
#             else:
#                 while True:
#                     if len(prev_data['candles']['data']) == 0 or counter == 2 or len(prev_data['candles']['data'][0]) == 0:
#                         break
#                     url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{sec[0]}/candles.json?from={prev_date_str}&till={prev_date_str}&interval=1&start={start-1}"
#                     prev_data = await get_prev(url, headers=headers, cookies=cookies)
#                     try:
#                         print(prev_data['candles']['data'][0])
#                         volumes.append(float(prev_data['candles']['data'][0][4]))
#                         print(f"{prev_data['candles']['data'][0][4]}")
#                         print(f"START: {start-1}")
#                         print(prev_date_str)
#                         # time.sleep(1)
#                     except Exception as e:
#                         counter+=1
#                     start += 1
#                 counter += 1
#         volumes_dict[f'{sec[0]}'] = sum(volumes) / start
#     return volumes_dict
            # prev_date = current_date - timedelta(days=day_counter)
            # prev_date_str = str(prev_date.strftime('%Y-%m-%d'))
            # url = f"https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{security}/candles.json?from={prev_date_str}&till={prev_date_str}&interval=1&start={start_from_for_prev_day}"
            # prev_data = await get_prev(url, headers=headers, cookies=cookies)
            # if len(prev_data['candles']['data']) == 0:
            #     day_counter += 1
            # else: 
            #     for candle_data in prev_data['candles']['data']:
            #         if cur_time in candle_data[6]:
            #             return candle_data
            #         else: 
            #             start_from_for_prev_day += 1
# CHECK VOLUMES BY FORMULA
# async def formula(security):
#     current_candle_data = await get_current_stock_volume(security)

#     if type(current_candle_data) is str:
#         return f"Ошибка при получении данных об акции formula {security}"
#     else: 
#         current_volume = current_candle_data[4]
#         return current_volume

def buyers_vs_sellers1(security):
    current_date = datetime.now(offset) # for test: - timedelta(days=1)
    today = current_date.strftime('%Y-%m-%d')
    url = f'https://iss.moex.com/iss/engines/stock/markets/shares/boards/TQBR/securities/{security}/candles.json?from={today}&till={today}&interval=10'
    response = re.get(url)
    data = response.json()['candles']['data']
    columns = response.json()['candles']['columns']
    df = pd.DataFrame(data, columns=columns)
    buyers = len(df[df['close'] > df['open']])
    sellers = len(df[df['close'] < df['open']])

    total = buyers + sellers
    if total > 0:
        buyers = round(buyers / total * 100)
        sellers = round(sellers / total * 100)
    else:
        buyers = 0
        sellers = 0

    return buyers, sellers



